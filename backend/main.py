import base64
import io
import json
import os
import re
from datetime import date
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, DefaultDict
import csv
import requests
from collections import defaultdict

from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, ListFlowable, ListItem

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from fastapi import HTTPException

app = FastAPI(title="Recuento diario de facturas")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_DIR = Path(__file__).parent
BASELINE_XLSX_PATH = DATA_DIR / "baseline.xlsx"
BASELINE_PDF_PATH = DATA_DIR / "baseline.pdf"
REPORTS_DIR = DATA_DIR / "reports"

# Added tipo y fecha para clasificar original/devolución y particionar por día.
HEADERS = [
    "TICKET DEVOLUCION",
    "TICKET FACTURA",
    "CAJA",
    "TIENDA",
    "VENDEDOR",
    "MONTO DEVUELTO",
    "MEDIO DE PAGO",
    "MOTIVO",
    "COMENTARIO",
    "TIPO",
    "FECHA",
]
COLUMN_KEYS = [
    "ticket_devolucion",
    "ticket_factura",
    "caja",
    "tienda",
    "vendedor",
    "monto_devuelto",
    "medio_pago",
    "motivo",
    "comentario",
    "tipo_documento",
    "fecha_operacion",
]


class EntriesPayload(BaseModel):
    entries: List[Dict[str, Any]]


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatPayload(BaseModel):
    messages: List[ChatMessage]


GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_CHAT_MODEL = os.getenv("GROQ_CHAT_MODEL", "llama-3.1-8b-instant")


def _baseline_exists() -> bool:
    return BASELINE_XLSX_PATH.exists() and BASELINE_XLSX_PATH.is_file() and BASELINE_XLSX_PATH.stat().st_size > 0


def _safe_float(value: str) -> float:
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace("$", "").replace("L", "").replace(" ", "")
    # handle thousands separators + decimal comma
    if s.count(",") > 0 and s.count(".") == 0:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def _load_baseline_schema() -> Dict[str, Any]:
    if not _baseline_exists():
        return {"available": False}

    try:
        wb = load_workbook(BASELINE_XLSX_PATH, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            header = []
            # heuristic: first row with >= 3 non-empty cells (up to 50 columns)
            for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
                values = [str(v).strip() for v in (row[:50] if row else []) if v is not None and str(v).strip()]
                if len(values) >= 3:
                    header = values
                    break
            sheets.append({"name": name, "header": header})
        return {"available": True, "sheets": sheets}
    except Exception:
        return {"available": True, "sheets": []}


def _baseline_pdf_exists() -> bool:
    return BASELINE_PDF_PATH.exists() and BASELINE_PDF_PATH.is_file() and BASELINE_PDF_PATH.stat().st_size > 0


_ES_MONTH_ABBR = {
    1: "Ene",
    2: "Feb",
    3: "Mar",
    4: "Abr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dic",
}


_ES_MONTH_NAME = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


def _parse_iso_date(value: str) -> datetime | None:
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except Exception:
        return None


def _format_currency(value: float) -> str:
    return f"${value:,.2f}"


def _format_day_es(dt: datetime) -> str:
    # ex: 01-Dic
    return f"{dt.day:02d}-{_ES_MONTH_ABBR.get(dt.month, str(dt.month))}"


def _mock_result() -> Dict[str, Any]:
    return {
        **{key: "" for key in COLUMN_KEYS},
        "tipo_documento": "devolucion",
        "fecha_operacion": date.today().isoformat(),
    }


def extract_data_from_image_bytes(image_bytes: bytes, filename: str) -> Dict[str, Any]:
    if not GROQ_API_KEY:
        return _mock_result()

    try:
        b64 = base64.b64encode(image_bytes).decode("utf-8")
        # Default to jpeg; Groq vision models accept data URLs
        data_url = f"data:image/jpeg;base64,{b64}"

        baseline_schema = _load_baseline_schema()
        baseline_hint = ""
        if baseline_schema.get("available") and baseline_schema.get("sheets"):
            sheet_names = [s.get("name", "") for s in baseline_schema["sheets"] if s.get("name")]
            baseline_hint = (
                "\nContexto adicional (baseline Excel cargado): "
                f"pestañas={sheet_names}. "
                "Usa este contexto solo para alinear criterios, pero responde siempre con las llaves solicitadas."
            )

        payload = {
            "model": "llama-3.2-11b-vision-preview",
            "temperature": 0.2,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Eres un asistente que extrae campos de tickets/facturas "
                        "y responde solo JSON válido con llaves: "
                        + ", ".join(COLUMN_KEYS)
                        + baseline_hint
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": (
                                "Extrae los campos de la imagen y responde JSON con claves en snake_case. "
                                "Si un valor no existe, deja cadena vacía. No agregues comentarios ni texto adicional."
                            ),
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": data_url, "detail": "low"},
                        },
                    ],
                },
            ],
            "response_format": {"type": "json_object"},
        }

        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        parsed = json.loads(content)
        return {key: str(parsed.get(key, "")) for key in COLUMN_KEYS}
    except Exception:
        # Fallback to mock on any failure
        return _mock_result()


def _call_groq_chat(messages: List[Dict[str, Any]]) -> str:
    if not GROQ_API_KEY:
        return "No hay GROQ_API_KEY configurada. (Respuesta mock)"

    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "model": GROQ_CHAT_MODEL,
            "temperature": 0.2,
            "messages": messages,
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def get_csv_path(fecha: str) -> Path:
    # Remove dashes to keep filename compact; default when missing fecha
    slug = fecha.replace("-", "") if fecha else "SIN_FECHA"
    return DATA_DIR / f"DEV_{slug}.csv"


def read_rows(path: Path) -> List[List[str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def ensure_structure(rows: List[List[str]]) -> List[List[str]]:
    # Ensure at least 6 blank rows before row 7
    while len(rows) < 6:
        rows.append([])

    # Row 7 text
    if len(rows) < 7:
        rows.append(["TIENDA CANAL DIGITAL T-45-63"])
    else:
        first_cell = rows[6][0] if rows[6] else ""
        rows[6] = [first_cell or "TIENDA CANAL DIGITAL T-45-63"]

    # Two spacer rows before headers to land on row 10
    while len(rows) < 9:
        rows.append([])

    if len(rows) < 10:
        rows.append(HEADERS)
    else:
        rows[9] = HEADERS

    return rows


def find_total_dev_index(rows: List[List[str]]) -> int:
    for idx, row in enumerate(rows):
        if row and row[0].strip().upper().startswith("TOTAL DEV"):
            return idx
    return -1


def write_rows(path: Path, rows: List[List[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def append_entries(entries: List[Dict[str, Any]]):
    grouped: DefaultDict[str, List[Dict[str, Any]]] = defaultdict(list)
    for entry in entries:
        fecha = str(entry.get("fecha_operacion", "")).strip()
        grouped[fecha].append(entry)

    saved_files = []
    for fecha, items in grouped.items():
        csv_path = get_csv_path(fecha)
        rows = ensure_structure(read_rows(csv_path))
        total_idx = find_total_dev_index(rows)
        insert_at = total_idx if total_idx != -1 else len(rows)

        data_rows = [[str(entry.get(key, "")) for key in COLUMN_KEYS] for entry in items]
        rows[insert_at:insert_at] = data_rows
        write_rows(csv_path, rows)
        saved_files.append(csv_path.name)

    return saved_files


def _iter_month_csv_files(year: int, month: int) -> List[Path]:
    # DEV_YYYYMMDD.csv
    prefix = f"DEV_{year:04d}{month:02d}"
    files = sorted(DATA_DIR.glob(f"{prefix}[0-9][0-9].csv"))
    return [p for p in files if p.is_file()]


def _read_entries_from_csv(path: Path) -> List[Dict[str, Any]]:
    rows = read_rows(path)
    if len(rows) < 11:
        return []
    data_rows = rows[10:]
    out: List[Dict[str, Any]] = []
    for r in data_rows:
        if not r:
            continue
        if r and str(r[0]).strip().upper().startswith("TOTAL DEV"):
            break
        # normalize length
        values = (r + [""] * len(COLUMN_KEYS))[: len(COLUMN_KEYS)]
        entry = {COLUMN_KEYS[i]: str(values[i]) for i in range(len(COLUMN_KEYS))}
        out.append(entry)
    return out


def _build_monthly_report_xlsx(year: int, month: int) -> bytes:
    entries: List[Dict[str, Any]] = []
    for f in _iter_month_csv_files(year, month):
        entries.extend(_read_entries_from_csv(f))

    total_amount = sum(_safe_float(e.get("monto_devuelto", "")) for e in entries)
    total_count = len(entries)

    by_day: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_tienda: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_motivo: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_medio: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    by_tipo: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})

    for e in entries:
        fecha = str(e.get("fecha_operacion", "")).strip() or "SIN_FECHA"
        tienda = str(e.get("tienda", "")).strip() or "(vacío)"
        motivo = str(e.get("motivo", "")).strip() or "(vacío)"
        medio = str(e.get("medio_pago", "")).strip() or "(vacío)"
        tipo = str(e.get("tipo_documento", "")).strip() or "(vacío)"
        amt = _safe_float(e.get("monto_devuelto", ""))

        by_day[fecha]["count"] += 1
        by_day[fecha]["amount"] += amt
        by_tienda[tienda]["count"] += 1
        by_tienda[tienda]["amount"] += amt
        by_motivo[motivo]["count"] += 1
        by_motivo[motivo]["amount"] += amt
        by_medio[medio]["count"] += 1
        by_medio[medio]["amount"] += amt
        by_tipo[tipo]["count"] += 1
        by_tipo[tipo]["amount"] += amt

    if _baseline_exists():
        wb = load_workbook(BASELINE_XLSX_PATH)
    else:
        wb = Workbook()

    # Ensure sheets
    resumen_name = "RESUMEN"
    detalle_name = "DETALLE"
    if resumen_name in wb.sheetnames:
        ws_res = wb[resumen_name]
        ws_res.delete_rows(1, ws_res.max_row)
    else:
        ws_res = wb.create_sheet(resumen_name)

    if detalle_name in wb.sheetnames:
        ws_det = wb[detalle_name]
        ws_det.delete_rows(1, ws_det.max_row)
    else:
        ws_det = wb.create_sheet(detalle_name)

    # Summary
    ws_res.append(["REPORTE MENSUAL DEVOLUCIONES"])
    ws_res.append(["Periodo", f"{year:04d}-{month:02d}"])
    ws_res.append(["Total registros", total_count])
    ws_res.append(["Total monto devuelto", total_amount])
    ws_res.append([])

    def _write_table(title: str, rows: List[List[Any]]):
        ws_res.append([title])
        for r in rows:
            ws_res.append(r)
        ws_res.append([])

    _write_table(
        "Por día",
        [["FECHA", "CANTIDAD", "MONTO"]]
        + [[k, int(v["count"]), float(v["amount"])] for k, v in sorted(by_day.items())],
    )
    _write_table(
        "Por tienda",
        [["TIENDA", "CANTIDAD", "MONTO"]]
        + [[k, int(v["count"]), float(v["amount"])] for k, v in sorted(by_tienda.items(), key=lambda kv: kv[1]["amount"], reverse=True)],
    )
    _write_table(
        "Por motivo",
        [["MOTIVO", "CANTIDAD", "MONTO"]]
        + [[k, int(v["count"]), float(v["amount"])] for k, v in sorted(by_motivo.items(), key=lambda kv: kv[1]["amount"], reverse=True)],
    )
    _write_table(
        "Por medio de pago",
        [["MEDIO DE PAGO", "CANTIDAD", "MONTO"]]
        + [[k, int(v["count"]), float(v["amount"])] for k, v in sorted(by_medio.items(), key=lambda kv: kv[1]["amount"], reverse=True)],
    )
    _write_table(
        "Por tipo",
        [["TIPO", "CANTIDAD", "MONTO"]]
        + [[k, int(v["count"]), float(v["amount"])] for k, v in sorted(by_tipo.items(), key=lambda kv: kv[1]["amount"], reverse=True)],
    )

    # Detail
    ws_det.append(HEADERS)
    for e in entries:
        ws_det.append([str(e.get(k, "")) for k in COLUMN_KEYS])

    # If workbook is empty default sheet, keep it but do not erase baseline sheets.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        # Leave as-is if baseline had it; otherwise remove auto-created sheet
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_monthly_report_pdf(year: int, month: int) -> bytes:
    entries: List[Dict[str, Any]] = []
    for f in _iter_month_csv_files(year, month):
        entries.extend(_read_entries_from_csv(f))

    # "Solo tickets reales": require at least one ticket id
    real_entries = [
        e
        for e in entries
        if str(e.get("ticket_devolucion", "")).strip() or str(e.get("ticket_factura", "")).strip()
    ]

    total_count = len(real_entries)
    total_amount = sum(_safe_float(e.get("monto_devuelto", "")) for e in real_entries)

    by_day: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for e in real_entries:
        fecha_raw = str(e.get("fecha_operacion", "")).strip()
        dt = _parse_iso_date(fecha_raw)
        key = _format_day_es(dt) if dt else (fecha_raw or "SIN_FECHA")
        by_day[key]["count"] += 1
        by_day[key]["amount"] += _safe_float(e.get("monto_devuelto", ""))

    # Peaks
    top_by_count = sorted(by_day.items(), key=lambda kv: (kv[1]["count"], kv[1]["amount"]), reverse=True)
    top_by_amount = sorted(by_day.items(), key=lambda kv: kv[1]["amount"], reverse=True)

    month_name = _ES_MONTH_NAME.get(month, str(month))

    def _pretty_day_phrase(day_key: str) -> str:
        # If day_key like 01-Dic and month is known, convert to "1 de diciembre"
        m = re.match(r"^(\d{2})-([A-Za-z]{3})$", day_key)
        if m:
            d = int(m.group(1))
            return f"{d} de {month_name}"
        # fallback: ISO date -> day of month
        dt = _parse_iso_date(day_key)
        if dt:
            return f"{dt.day} de {_ES_MONTH_NAME.get(dt.month, str(dt.month))}"
        return day_key

    peak_lines = []
    if top_by_count:
        for item in top_by_count[:2]:
            day_key, stats = item
            peak_lines.append(f"{_pretty_day_phrase(day_key)}: {int(stats['count'])} tickets.")

    max_amount_line = None
    if top_by_amount:
        day_key, stats = top_by_amount[0]
        max_amount_line = (
            f"Día con mayor impacto económico: {_pretty_day_phrase(day_key)} "
            f"(Monto: {_format_currency(float(stats['amount']))})."
        )

    # Hallazgos: refact + logística
    refact_count = 0
    refact_amount = 0.0
    inv_cedi_count = 0
    for e in real_entries:
        motivo = str(e.get("motivo", "")).lower()
        comentario = str(e.get("comentario", "")).lower()
        amt = _safe_float(e.get("monto_devuelto", ""))
        if "refact" in motivo or "refact" in comentario:
            refact_count += 1
            refact_amount += amt
        if "cedi" in comentario or "inventario" in comentario:
            inv_cedi_count += 1

    # Top vendedores
    by_vendor: DefaultDict[str, Dict[str, float]] = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for e in real_entries:
        vendor = str(e.get("vendedor", "")).strip() or "(vacío)"
        by_vendor[vendor]["count"] += 1
        by_vendor[vendor]["amount"] += _safe_float(e.get("monto_devuelto", ""))
    top_vendors = sorted(by_vendor.items(), key=lambda kv: kv[1]["amount"], reverse=True)[:5]

    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    body = styles["BodyText"]
    body.spaceAfter = 8
    body.leading = 14
    bullet_style = ParagraphStyle(
        "bullet",
        parent=styles["BodyText"],
        leftIndent=18,
        leading=14,
        spaceAfter=4,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.8 * inch,
        rightMargin=0.8 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        title=f"Reporte {year:04d}-{month:02d}",
    )

    story: List[Any] = []

    # 1) Resumen ejecutivo
    story.append(Paragraph("1. RESUMEN EJECUTIVO", h1))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Reporte de Devoluciones (Solo Tickets Reales)", body))
    bullets = [
        f"Total de Devoluciones procesadas: {total_count} tickets.",
        f"Monto Total Acumulado: {_format_currency(total_amount)}.",
        "Días con mayor volumen de transacciones:",
    ]
    story.append(ListFlowable([ListItem(Paragraph(x, bullet_style)) for x in bullets[:2]], bulletType="bullet"))
    story.append(ListFlowable([ListItem(Paragraph(bullets[2], bullet_style))], bulletType="bullet"))
    if peak_lines:
        sub_style = ParagraphStyle(
            "sub",
            parent=styles["BodyText"],
            leftIndent=34,
            leading=14,
            spaceAfter=2,
        )
        for line in peak_lines[:2]:
            story.append(Paragraph(f"o {line}", sub_style))
    if max_amount_line:
        story.append(ListFlowable([ListItem(Paragraph(max_amount_line, bullet_style))], bulletType="bullet"))

    story.append(Spacer(1, 10))

    # 2) Volumen diario
    story.append(Paragraph("2. ANÁLISIS DE VOLUMEN DIARIO", h1))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Desglose de Tickets por Fecha", h2))
    story.append(
        Paragraph(
            "Este conteo representa cuántas devoluciones (filas de tickets) ocurrieron cada día:",
            body,
        )
    )

    day_rows = [["Fecha", "Cantidad de Tickets", "Monto del Día"]]
    for day_key, stats in sorted(by_day.items()):
        day_rows.append([day_key, int(stats["count"]), _format_currency(float(stats["amount"]))])
    tbl = Table(day_rows, colWidths=[1.2 * inch, 1.8 * inch, 2.0 * inch])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 12))

    # 3) Hallazgos clave
    story.append(Paragraph("3. Hallazgos Clave", h1))
    story.append(Spacer(1, 6))
    hk = []
    if top_by_count:
        d1, s1 = top_by_count[0]
        d2, s2 = (top_by_count[1] if len(top_by_count) > 1 else (None, None))
        peaks_phrase = (
            f"Picos Operativos: Los días {_pretty_day_phrase(d1)} y {_pretty_day_phrase(d2)} fueron los de mayor carga administrativa "
            f"con {int(s1['count'])} tickets" + (f" y {int(s2['count'])} tickets" if s2 else "") + "."
            if d2
            else f"Picos Operativos: El día {_pretty_day_phrase(d1)} fue el de mayor carga administrativa con {int(s1['count'])} tickets."
        )
        if top_by_amount:
            da, sa = top_by_amount[0]
            peaks_phrase += (
                f" El {_pretty_day_phrase(da)} registró el mayor monto del mes "
                f"({_format_currency(float(sa['amount']))}), lo que sugiere devoluciones de mayor valor."
            )
        hk.append(peaks_phrase)

    if refact_count > 0:
        hk.append(
            f"Eficiencia de Refacturación: Se identificaron {refact_count} registros con indicios de refacturación "
            f"(monto asociado: {_format_currency(refact_amount)}). Conviene validar si existe recompra en el mismo movimiento o posterior."
        )
    else:
        hk.append(
            "Eficiencia de Refacturación: No se detectaron referencias claras a refacturación en motivo/comentario; "
            "si este proceso aplica, conviene estandarizar cómo se registra en el ticket para poder medirlo."
        )

    if inv_cedi_count > 0:
        hk.append(
            f"Problema Logístico: {inv_cedi_count} comentarios mencionan inventario/CEDI, lo que sugiere impacto directo de abastecimiento en la experiencia del cliente."
        )
    else:
        hk.append(
            "Problema Logístico: No se detectaron menciones frecuentes de inventario/CEDI en comentarios; "
            "si es un problema recurrente, conviene capturarlo con un motivo estandarizado."
        )

    story.append(ListFlowable([ListItem(Paragraph(x, bullet_style)) for x in hk], bulletType='bullet'))
    story.append(Spacer(1, 12))

    # 4) Top vendedores
    story.append(Paragraph("4. Top 5 Vendedores por Monto de Devolución", h1))
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "Identificar a los vendedores con mayores montos ayuda a detectar necesidades de capacitación en el proceso de venta o facturación.",
            body,
        )
    )
    vendor_lines = []
    for i, (vendor, stats) in enumerate(top_vendors, start=1):
        vendor_lines.append(
            Paragraph(
                f"{i}. {vendor}: {_format_currency(float(stats['amount']))} ({int(stats['count'])} tickets)",
                body,
            )
        )
    story.extend(vendor_lines if vendor_lines else [Paragraph("Sin datos suficientes.", body)])

    doc.build(story)
    return buf.getvalue()


@app.post("/upload")
async def upload(files: List[UploadFile] = File(...)):
    results = []
    for file in files:
        content = await file.read()
        results.append(extract_data_from_image_bytes(content, file.filename))
    return {"results": results}


@app.post("/baseline")
async def baseline(file: UploadFile = File(...)):
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".xls", ".pdf")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx, .xls o .pdf")
    content = await file.read()
    if name.endswith((".xlsx", ".xls")):
        BASELINE_XLSX_PATH.write_bytes(content)
        return {"status": "baseline guardada", "file": BASELINE_XLSX_PATH.name}
    BASELINE_PDF_PATH.write_bytes(content)
    return {"status": "baseline guardada", "file": BASELINE_PDF_PATH.name}


@app.get("/baseline/status")
async def baseline_status():
    xlsx_ok = _baseline_exists()
    pdf_ok = _baseline_pdf_exists()
    payload: Dict[str, Any] = {
        "available": bool(xlsx_ok or pdf_ok),
        "xlsx": None,
        "pdf": None,
    }
    if xlsx_ok:
        st = BASELINE_XLSX_PATH.stat()
        payload["xlsx"] = {
            "file": BASELINE_XLSX_PATH.name,
            "bytes": st.st_size,
            "modified": int(st.st_mtime),
        }
    if pdf_ok:
        st = BASELINE_PDF_PATH.stat()
        payload["pdf"] = {
            "file": BASELINE_PDF_PATH.name,
            "bytes": st.st_size,
            "modified": int(st.st_mtime),
        }
    return payload


@app.get("/baseline/schema")
async def baseline_schema():
    return _load_baseline_schema()


@app.get("/report/monthly")
async def report_monthly(year: int, month: int):
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month debe ser 1-12")
    content = _build_monthly_report_xlsx(year, month)
    filename = f"reporte_{year:04d}_{month:02d}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/report/monthly/pdf")
async def report_monthly_pdf(year: int, month: int):
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month debe ser 1-12")
    content = _build_monthly_report_pdf(year, month)
    filename = f"reporte_{year:04d}_{month:02d}.pdf"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/save")
async def save(payload: EntriesPayload):
    files = append_entries(payload.entries)
    return {"status": "guardado", "files": files}


@app.post("/chat")
async def chat(payload: ChatPayload):
    # Minimal chat endpoint for optional UI widget.
    incoming = payload.messages or []

    baseline_schema = _load_baseline_schema()
    baseline_note = ""
    if baseline_schema.get("available"):
        sheet_names = [s.get("name") for s in baseline_schema.get("sheets", []) if s.get("name")]
        if sheet_names:
            baseline_note = f"\nBaseline cargado (pestañas): {sheet_names}."

    system = (
        "Eres un asistente para el sistema de recuento diario de facturas/devoluciones. "
        "Responde en español, de forma breve y accionable. "
        "Si te preguntan por pasos, da instrucciones concretas."
        + baseline_note
    )

    # Keep only last N user/assistant messages to control prompt size.
    trimmed = incoming[-20:]
    messages = [{"role": "system", "content": system}] + [m.model_dump() for m in trimmed]

    try:
        answer = _call_groq_chat(messages)
        return {"reply": answer}
    except Exception:
        return {"reply": "No pude generar respuesta en este momento. Intenta de nuevo."}
